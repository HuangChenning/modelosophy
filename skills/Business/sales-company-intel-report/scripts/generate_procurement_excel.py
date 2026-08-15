#!/usr/bin/env python3
"""Export the it_bidding section of data.json to a filterable Excel workbook.

Produces two sheets:
  - "招投标明细": one row per record in it_bidding.records
  - "供应商汇总": one row per entry in it_bidding.vendor_summary

Usage:
    python generate_procurement_excel.py --data data.json --output 招投标明细.xlsx
"""
from __future__ import annotations

import argparse
import json
import logging
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit(
        "openpyxl is required: pip install openpyxl --break-system-packages"
    ) from exc

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1A2233", end_color="1A2233", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

RECORD_COLUMNS = [
    ("date", "时间", 12),
    ("project_name", "项目名称", 42),
    ("amount_display", "金额（原始表述）", 18),
    ("amount", "金额（数值，元）", 16),
    ("category", "类别", 18),
    ("vendor", "中标供应商", 26),
    ("confidence", "置信度", 10),
    ("source_name", "来源", 20),
    ("source_url", "来源链接", 30),
]

VENDOR_COLUMNS = [
    ("vendor", "供应商", 26),
    ("win_count", "中标次数", 10),
    ("total_amount", "中标金额合计（元）", 18),
    ("share_pct", "金额份额（%）", 14),
    ("categories", "项目类别", 30),
]


def load_data(data_path: Path) -> dict[str, Any]:
    if not data_path.exists():
        raise FileNotFoundError(f"data file not found: {data_path}")
    with data_path.open("r", encoding="utf-8") as f:
        try:
            data = json.load(f)
        except json.JSONDecodeError as exc:
            raise ValueError(f"invalid JSON in {data_path}: {exc}") from exc
    return data


def write_sheet(ws: Worksheet, columns: list[tuple[str, str, int]], rows: list[dict[str, Any]]) -> None:
    for col_idx, (_, header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _, _) in enumerate(columns, start=1):
            value = row.get(key)
            if key == "categories" and isinstance(value, list):
                value = "、".join(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    if not rows:
        ws.cell(row=2, column=1, value="未检索到相关记录")


def build_workbook(data: dict[str, Any]) -> Workbook:
    it_bidding = data.get("it_bidding", {}) or {}
    records = it_bidding.get("records") or []
    vendor_summary = it_bidding.get("vendor_summary") or []
    vendor_summary = sorted(vendor_summary, key=lambda v: v.get("total_amount") or 0, reverse=True)

    wb = Workbook()
    ws_records = wb.active
    ws_records.title = "招投标明细"
    write_sheet(ws_records, RECORD_COLUMNS, records)

    ws_vendors = wb.create_sheet("供应商汇总")
    write_sheet(ws_vendors, VENDOR_COLUMNS, vendor_summary)

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--data", required=True, type=Path, help="Path to data.json")
    parser.add_argument("--output", required=True, type=Path, help="Path to write the .xlsx file")
    args = parser.parse_args()

    data = load_data(args.data)
    records = (data.get("it_bidding", {}) or {}).get("records") or []
    if not records:
        logger.warning(
            "it_bidding.records is empty — writing a workbook with header rows only. "
            "Confirm this is expected before sharing the file."
        )

    wb = build_workbook(data)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    logger.info("Excel written to %s (%d bidding records)", args.output, len(records))


if __name__ == "__main__":
    main()
